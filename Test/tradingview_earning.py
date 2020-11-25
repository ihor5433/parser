from selenium import webdriver
from time import sleep
from bs4 import BeautifulSoup as bs
import pandas as pd
from datetime import datetime,timedelta
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter


# wb = load_workbook('1.xlsx')
# sheet = wb["Sheet1"]
# amountOfColumns = sheet.max_column
# amountOfRows = sheet.max_row
def parse():
    chromedriver = 'C:\\Users\idmytrenko\Documents\chromedriver.exe'
    options = webdriver.ChromeOptions()
    browser = webdriver.Chrome(executable_path=chromedriver,options=options)

    browser.get('https://ru.tradingview.com/markets/stocks-usa/earnings/')

    elements = browser.find_element_by_xpath('//*[@id="js-screener-container"]/div[2]/div[6]/div/div/div[1]/div').click()
    sleep(1)
    #elements = browser.find_element_by_class_name('tv-load-more__btn').click()
    sleep(1)
    html = browser.page_source
    return  html

def tradingview_earning(html):
    """
    Выкачивает отчеты за сегодня с TradingView
    """
    # soup = convert_html(url)
    soup = bs(html,features="lxml")
    tables_row = soup.find_all("tr")
    tags = soup.find_all(class_="tv-screener__symbol")
    texts = [i.text for i in tags if i.text]
    res = []
    c = 0
    row = []
    for tr in tables_row:
        td = tr.find_all("td")
        row = [tr.text for tr in td[:-3] if tr.text]
        if row:
            res.append(row)
    for i in res:
        n = texts[c]
        del i[0]
        i.insert(0, n)
        c += 1
        if c == len(res):
            break
    columns = ["Тикер", "Рын.кап", "Ожидание EPS", "EPS", "Отклонение", "Отклонение в %", 'Прогноз выручки', "Выручка"]
    df = pd.DataFrame(res,columns=columns)
    print(df.to_markdown())
    df.to_csv(f'{datetime.now().date() - timedelta(days=1)}.csv',index=False,encoding='utf-8')
    df.to_excel(f'{datetime.now().date() - timedelta(days=1)}.xlsx',index=False)

tradingview_earning(parse())

# for i in range(amountOfColumns):
#     for k in range(amountOfRows):
#         cell = str(sheet[get_column_letter(i+1)+str(k+1)].value)

#         if cell.find('.') != -1:
#             dot = cell.find('.')



input()

